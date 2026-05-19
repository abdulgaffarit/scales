#!/usr/bin/env python3
"""
Download and parse the latest BLS OES May 2025 national data
into the format needed by build.py (data/jobs.csv)
"""

import requests
import zipfile
import io
import csv
import os
from pathlib import Path
from openpyxl import load_workbook

DATA_DIR = Path(__file__).parent / "data"
NATIONAL_URL = "https://www.bls.gov/oes/special-requests/oesm25nat.zip"

# Mapping of SOC major groups to our categories
SOC_CATEGORY_MAP = {
    "11": "Management",
    "13": "Business",
    "15": "Technology",
    "17": "Engineering",
    "19": "Science",
    "21": "Social Services",
    "23": "Legal",
    "25": "Education",
    "27": "Creative",
    "29": "Healthcare",
    "31": "Healthcare",
    "33": "Public Safety",
    "35": "Hospitality",
    "37": "Personal Services",
    "39": "Personal Services",
    "41": "Sales",
    "43": "Business",
    "45": "Agriculture",
    "47": "Construction",
    "49": "Skilled Trades",
    "51": "Manufacturing",
    "53": "Transportation",
}

# Demand level based on employment count
def get_demand(employment):
    if employment >= 500000:
        return "Very High"
    elif employment >= 200000:
        return "High"
    elif employment >= 50000:
        return "Medium"
    else:
        return "Medium"

def slugify(title):
    """Convert job title to URL slug"""
    import re
    s = title.lower().strip()
    s = re.sub(r'[^a-z0-9\s-]', '', s)
    s = re.sub(r'[\s]+', '-', s)
    s = re.sub(r'-+', '-', s)
    return s.strip('-')


def download_and_extract(url):
    """Download zip and extract the Excel file, or use local file"""
    local_xlsx = Path(__file__).parent / "national_M2025_dl.xlsx"
    local_xlsx2 = Path(__file__).parent / "national-M2025-dl.xlsx"
    local_zip = Path(__file__).parent / "oesm25nat.zip"

    # Check if already downloaded manually
    for f in [local_xlsx, local_xlsx2]:
        if f.exists():
            print(f"📂 Using local file: {f}")
            return f.read_bytes()

    if local_zip.exists():
        print(f"📂 Using local zip: {local_zip}")
        z = zipfile.ZipFile(local_zip)
        xlsx_files = [f for f in z.namelist() if f.endswith('.xlsx')]
        if xlsx_files:
            print(f"   Extracting: {xlsx_files[0]}")
            return z.read(xlsx_files[0])

    # Try downloading (may fail due to BLS bot protection)
    print(f"📥 Attempting download from {url}...")
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
        "Accept": "application/zip, application/octet-stream, */*",
        "Referer": "https://www.bls.gov/oes/tables.htm",
    }
    try:
        resp = requests.get(url, timeout=60, headers=headers)
        resp.raise_for_status()
        print(f"   Downloaded {len(resp.content) / 1024 / 1024:.1f} MB")
    except requests.exceptions.HTTPError as e:
        print(f"\n❌ BLS blocked automated download (403 Forbidden).")
        print(f"   BLS requires manual browser download to prevent bot access.")
        print(f"\n📋 MANUAL DOWNLOAD INSTRUCTIONS:")
        print(f"   1. Open this URL in your browser:")
        print(f"      {url}")
        print(f"   2. Save the zip file to this project folder")
        print(f"   3. Run this script again")
        print(f"\n   Alternative: Download 'All data' from:")
        print(f"      https://www.bls.gov/oes/tables.htm")
        raise SystemExit(1)

    z = zipfile.ZipFile(io.BytesIO(resp.content))
    xlsx_files = [f for f in z.namelist() if f.endswith('.xlsx')]
    if not xlsx_files:
        raise Exception(f"No xlsx files found in zip. Contents: {z.namelist()}")

    print(f"   Extracting: {xlsx_files[0]}")
    return z.read(xlsx_files[0])

    z = zipfile.ZipFile(io.BytesIO(resp.content))
    xlsx_files = [f for f in z.namelist() if f.endswith('.xlsx')]
    if not xlsx_files:
        raise Exception(f"No xlsx files found in zip. Contents: {z.namelist()}")

    print(f"   Extracting: {xlsx_files[0]}")
    return z.read(xlsx_files[0])


def parse_national_data(xlsx_bytes):
    """Parse the BLS national XLSX into job records"""
    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)

    # Find the right sheet - usually named something like "national_M2025_dl"
    sheet_names = wb.sheetnames
    print(f"   Sheets found: {sheet_names}")

    # Use the first sheet
    ws = wb[sheet_names[0]]

    # Read header row to find column indices
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    header = [str(h).strip().upper() if h else "" for h in header]
    print(f"   Columns: {header[:20]}...")

    # Map column names to indices
    col_map = {}
    for i, h in enumerate(header):
        col_map[h] = i

    # Verify key columns exist
    needed = ["OCC_CODE", "OCC_TITLE", "O_GROUP", "TOT_EMP", "A_MEAN", "A_MEDIAN", "A_PCT10", "A_PCT90"]
    for col in needed:
        if col not in col_map:
            print(f"   ⚠️  Column '{col}' not found! Available: {list(col_map.keys())[:20]}")

    jobs = []
    skipped = 0
    for row in rows:
        if not row or len(row) < 10:
            continue

        # Get values
        occ_code = str(row[col_map.get("OCC_CODE", 0)] or "").strip()
        occ_title = str(row[col_map.get("OCC_TITLE", 1)] or "").strip()
        o_group = str(row[col_map.get("O_GROUP", 0)] or "").strip().lower()

        # Only keep "detailed" level occupations (most specific)
        if o_group != "detailed":
            continue

        # Skip "All Occupations" and codes ending in -0000
        if not occ_code or occ_code.endswith("-0000") or "All Occupations" in occ_title:
            continue

        # Only keep standard SOC codes (XX-XXXX format)
        if len(occ_code) != 7 or occ_code[2] != '-':
            continue

        # Get numeric values
        def safe_int(val):
            if val is None:
                return 0
            s = str(val).replace(",", "").replace("$", "").strip()
            if s in ("*", "**", "#", "N/A", "", "nan", "None"):
                return 0
            try:
                return int(float(s))
            except (ValueError, TypeError):
                return 0

        employment = safe_int(row[col_map.get("TOT_EMP", 0)])
        avg = safe_int(row[col_map.get("A_MEAN", 0)])
        median = safe_int(row[col_map.get("A_MEDIAN", 0)])
        p10 = safe_int(row[col_map.get("A_PCT10", 0)])
        p25 = safe_int(row[col_map.get("A_PCT25", 0)])
        p75 = safe_int(row[col_map.get("A_PCT75", 0)])
        p90 = safe_int(row[col_map.get("A_PCT90", 0)])

        # Skip if no meaningful salary data
        if avg == 0 and median == 0:
            skipped += 1
            continue
        if employment == 0:
            skipped += 1
            continue

        # Use median as fallback for avg and vice versa
        if avg == 0:
            avg = median
        if median == 0:
            median = avg

        # Skip very low salary occupations (likely part-time or data issues)
        if avg < 25000:
            skipped += 1
            continue

        # Determine category from SOC code prefix
        soc_prefix = occ_code[:2]
        category = SOC_CATEGORY_MAP.get(soc_prefix, "Business")

        # Calculate YoY growth estimate based on category
        growth_map = {
            "Technology": 5.5, "Healthcare": 4.2, "Engineering": 3.2,
            "Science": 3.8, "Management": 3.5, "Legal": 3.0,
            "Education": 2.2, "Business": 3.5, "Skilled Trades": 3.2,
            "Construction": 3.5, "Sales": 2.8, "Creative": 3.0,
            "Transportation": 2.5, "Hospitality": 3.0, "Personal Services": 3.0,
            "Public Safety": 2.5, "Social Services": 4.0,
            "Manufacturing": 2.5, "Agriculture": 2.0,
        }
        yoy = growth_map.get(category, 3.0)

        demand = get_demand(employment)

        jobs.append({
            "soc": occ_code,
            "job_slug": slugify(occ_title),
            "job_title": occ_title,
            "category": category,
            "national_avg": avg,
            "national_median": median,
            "national_low": p10,
            "national_high": p90,
            "yoy_growth": yoy,
            "demand": demand,
            "employment": employment,
        })

    wb.close()
    print(f"   Skipped {skipped} rows (no data, too low salary, or non-detailed)")
    return jobs


def write_jobs_csv(jobs):
    """Write jobs to data/jobs.csv"""
    DATA_DIR.mkdir(exist_ok=True)
    output_path = DATA_DIR / "jobs.csv"

    # Sort by category then by avg salary descending
    jobs.sort(key=lambda j: (j["category"], -j["national_avg"]))

    fieldnames = ["job_slug", "job_title", "category", "national_avg", "national_median",
                  "national_low", "national_high", "yoy_growth", "demand", "employment"]

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for job in jobs:
            writer.writerow({k: job[k] for k in fieldnames})

    print(f"\n✅ Written {len(jobs)} jobs to {output_path}")
    return output_path


def main():
    print("🚀 BLS OES May 2025 Data Updater")
    print("=" * 50)

    # Download national data
    xlsx_bytes = download_and_extract(NATIONAL_URL)

    # Parse it
    print("\n📊 Parsing national occupation data...")
    jobs = parse_national_data(xlsx_bytes)
    print(f"   Found {len(jobs)} detailed occupations with salary data")

    # Show category breakdown
    from collections import Counter
    cats = Counter(j["category"] for j in jobs)
    print("\n📂 Category breakdown:")
    for cat, count in sorted(cats.items(), key=lambda x: -x[1]):
        print(f"   {cat}: {count} jobs")

    # Show top 10 highest paying
    top = sorted(jobs, key=lambda j: -j["national_avg"])[:10]
    print("\n💰 Top 10 highest paying:")
    for j in top:
        print(f"   ${j['national_avg']:>9,}  {j['job_title']}")

    # Write to CSV
    write_jobs_csv(jobs)

    print("\n🎉 Data update complete!")
    print("   Run 'python build.py' to rebuild the site with new data.")


if __name__ == "__main__":
    main()
