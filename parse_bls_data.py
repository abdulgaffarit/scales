#!/usr/bin/env python3
"""
Parse BLS OES May 2025 data directly from the XLSX file.
Place the file 'national-M2025-dl.xlsx' or 'national_M2025_dl.xlsx' in this folder,
then run: python parse_bls_data.py

If you don't have the file yet, download it from:
https://www.bls.gov/oes/special-requests/oesm25nat.zip
"""
import csv
import re
from pathlib import Path

DATA_DIR = Path(__file__).parent / "data"

# SOC prefix to category mapping
SOC_CATEGORY_MAP = {
    "11": "Management", "13": "Business", "15": "Technology",
    "17": "Engineering", "19": "Science", "21": "Social Services",
    "23": "Legal", "25": "Education", "27": "Creative",
    "29": "Healthcare", "31": "Healthcare", "33": "Public Safety",
    "35": "Hospitality", "37": "Personal Services", "39": "Personal Services",
    "41": "Sales", "43": "Business", "45": "Agriculture",
    "47": "Construction", "49": "Skilled Trades", "51": "Manufacturing",
    "53": "Transportation",
}

def get_demand(emp):
    if emp >= 500000: return "Very High"
    elif emp >= 200000: return "High"
    elif emp >= 50000: return "Medium"
    else: return "Medium"

def slugify(title):
    s = title.lower().strip()
    s = re.sub(r'[^a-z0-9\s-]', '', s)
    s = re.sub(r'[\s]+', '-', s)
    s = re.sub(r'-+', '-', s)
    return s.strip('-')


def safe_int(val):
    if val is None: return 0
    s = str(val).replace(",", "").replace("$", "").strip()
    if s in ("*", "**", "#", "N/A", "", "nan", "None", "TRUE"): return 0
    try: return int(float(s))
    except: return 0

def parse_xlsx():
    """Try to parse from XLSX file using openpyxl"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("openpyxl not installed. Run: pip install openpyxl")
        return None

    # Find the file
    base = Path(__file__).parent
    candidates = [
        base / "national-M2025-dl.xlsx",
        base / "national_M2025_dl.xlsx",
        base / "oesm25nat" / "national_M2025_dl.xlsx",
    ]
    xlsx_path = None
    for c in candidates:
        if c.exists():
            xlsx_path = c
            break

    if not xlsx_path:
        print("❌ Could not find the BLS Excel file.")
        print("   Please place 'national-M2025-dl.xlsx' in this folder.")
        print("   Download from: https://www.bls.gov/oes/special-requests/oesm25nat.zip")
        return None

    print(f"📂 Reading: {xlsx_path}")
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip().upper() if h else "" for h in next(rows)]
    col = {h: i for i, h in enumerate(header)}
    print(f"   Columns found: {len(header)}")

    jobs = []
    for row in rows:
        if not row or len(row) < 15: continue
        o_group = str(row[col.get("O_GROUP", 0)] or "").strip().lower()
        if o_group != "detailed": continue

        occ_code = str(row[col.get("OCC_CODE", 0)] or "").strip()
        occ_title = str(row[col.get("OCC_TITLE", 0)] or "").strip()
        if not occ_code or occ_code.endswith("-0000"): continue
        if len(occ_code) != 7 or occ_code[2] != '-': continue

        emp = safe_int(row[col.get("TOT_EMP", 0)])
        avg = safe_int(row[col.get("A_MEAN", 0)])
        med = safe_int(row[col.get("A_MEDIAN", 0)])
        p10 = safe_int(row[col.get("A_PCT10", 0)])
        p90 = safe_int(row[col.get("A_PCT90", 0)])

        if (avg == 0 and med == 0) or emp == 0: continue
        if avg == 0: avg = med
        if med == 0: med = avg
        if avg < 25000: continue

        soc_prefix = occ_code[:2]
        category = SOC_CATEGORY_MAP.get(soc_prefix, "Business")
        growth_map = {
            "Technology": 5.5, "Healthcare": 4.2, "Engineering": 3.2,
            "Science": 3.8, "Management": 3.5, "Legal": 3.0,
            "Education": 2.2, "Business": 3.5, "Skilled Trades": 3.2,
            "Construction": 3.5, "Sales": 2.8, "Creative": 3.0,
            "Transportation": 2.5, "Hospitality": 3.0, "Personal Services": 3.0,
            "Public Safety": 2.5, "Social Services": 4.0,
            "Manufacturing": 2.5, "Agriculture": 2.0,
        }
        jobs.append({
            "job_slug": slugify(occ_title),
            "job_title": occ_title,
            "category": category,
            "national_avg": avg,
            "national_median": med,
            "national_low": p10 if p10 > 0 else round(avg * 0.55),
            "national_high": p90 if p90 > 0 else round(avg * 1.55),
            "yoy_growth": growth_map.get(category, 3.0),
            "demand": get_demand(emp),
            "employment": emp,
        })

    wb.close()
    return jobs

def write_csv(jobs):
    DATA_DIR.mkdir(exist_ok=True)
    out = DATA_DIR / "jobs.csv"
    jobs.sort(key=lambda j: (j["category"], -j["national_avg"]))
    fields = ["job_slug","job_title","category","national_avg","national_median",
              "national_low","national_high","yoy_growth","demand","employment"]
    with open(out, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for j in jobs:
            w.writerow({k: j[k] for k in fields})
    print(f"\n✅ Written {len(jobs)} jobs to {out}")

if __name__ == "__main__":
    print("🚀 BLS OES May 2025 → jobs.csv converter")
    print("=" * 50)
    jobs = parse_xlsx()
    if jobs:
        from collections import Counter
        cats = Counter(j["category"] for j in jobs)
        print(f"\n📊 Found {len(jobs)} occupations across {len(cats)} categories")
        for cat, n in sorted(cats.items(), key=lambda x: -x[1])[:10]:
            print(f"   {cat}: {n}")
        top = sorted(jobs, key=lambda j: -j["national_avg"])[:5]
        print(f"\n💰 Top 5 highest paying:")
        for j in top:
            print(f"   ${j['national_avg']:>9,}  {j['job_title']}")
        write_csv(jobs)
        print("\n🎉 Done! Run 'python build.py' to rebuild the site.")
